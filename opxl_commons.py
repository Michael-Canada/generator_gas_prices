from datetime import datetime
from typing import Any, Iterable, Optional, NamedTuple, Sequence, SupportsFloat, cast, Tuple, Union

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import (
    TYPE_NUMERIC,
    TYPE_STRING,
    Cell,
)

# from sourcing.utils.date_utils import NaiveDateTime
from date_utils import NaiveDateTime


_EXPECTED_TYPES = {
    TYPE_NUMERIC,
    TYPE_STRING,
}


def get_row(
    worksheet: Worksheet, row_zero_indexed: int, ncols: int
) -> Tuple[Cell, ...]:
    row_one_indexed = row_zero_indexed + 1

    row = list(
        worksheet.iter_rows(
            min_row=row_one_indexed, max_row=row_one_indexed, max_col=ncols
        )
    )
    assert len(row) == 1

    return row[0]


def iter_rows(
    worksheet: Worksheet,
    min_row_zero_indexed: int,
    max_row_zero_indexed: int,
    ncols: int,
) -> Iterable[Tuple[Cell, ...]]:

    min_row_one_indexed = min_row_zero_indexed + 1
    max_row_one_indexed = max_row_zero_indexed + 1

    yield from worksheet.iter_rows(
        min_row=min_row_one_indexed, max_row=max_row_one_indexed, max_col=ncols
    )


class UnsupportedCellType(Exception):
    pass


def _xl_parse(
    cell: Cell,
    expected_type: str,
    allow_empty: bool = False,
    empty_text_as_null: bool = False,
) -> Any:

    assert expected_type in _EXPECTED_TYPES

    if cell.value is None:
        assert allow_empty, f"Got empty when not allowed for cell {cell}"
        return None

    if expected_type != cell.data_type:
        # assert (
        #     empty_text_as_null
        #     and cell.data_type == TYPE_STRING
        #     and isinstance(cell.value, str)
        #     and cell.value.strip() == ""
        # ), f"Expected type {expected_type} for cell {str(cell)}"

        return None

    if cell.data_type == TYPE_STRING:
        assert isinstance(cell.value, str)

        value = cell.value.strip()
        if value == "":
            assert allow_empty, "Got empty string when allow_empty is False"
            return None

        return value
    elif cell.data_type == TYPE_NUMERIC:
        return float(cast(SupportsFloat, cell.value))
    else:
        raise UnsupportedCellType(f"Cannot parse cell type: {cell.data_type}")


def is_none(cell: Cell) -> bool:
    if cell.value is None:
        return True
    elif cell.data_type == TYPE_STRING:
        value = _xl_parse(cell, TYPE_STRING, allow_empty=True, empty_text_as_null=True)
        return value is None
    else:
        return False


def is_number(cell: Cell) -> bool:
    return cell.data_type == TYPE_NUMERIC and cell.value is not None


def ensure_int(cell: Cell) -> int:
    value: float = _xl_parse(cell, TYPE_NUMERIC)

    return _float_to_int(value)


def _float_to_int(value: float) -> int:
    assert int(value) == value, "Value provided has data after the decimal point."
    return int(value)


def ensure_int_or_none(cell: Cell) -> Optional[int]:
    value: Optional[float] = _xl_parse(
        cell, TYPE_NUMERIC, allow_empty=True, empty_text_as_null=True
    )

    if value is None:
        return None

    return _float_to_int(value)


def non_empty_string(cell: Cell, float_to_str: bool = False) -> str:
    if float_to_str and cell.data_type == TYPE_NUMERIC:
        value = _xl_parse(cell, TYPE_NUMERIC)

        return str(value)

    value = _xl_parse(cell, TYPE_STRING)

    return value


def str_parse_or_none(cell: Cell) -> Optional[str]:
    value = _xl_parse(cell, TYPE_STRING, allow_empty=True, empty_text_as_null=True)

    return value


def float_parse(cell: Cell) -> float:
    value = _xl_parse(cell, TYPE_NUMERIC)

    return value


def float_parse_or_none(cell: Cell) -> Optional[float]:
    value = _xl_parse(cell, TYPE_NUMERIC, allow_empty=True, empty_text_as_null=True)

    return value


def parse_date(cell: Cell) -> NaiveDateTime:
    dt = cell.value

    assert isinstance(dt, datetime)
    assert dt.tzinfo is None

    return NaiveDateTime(dt)


class WrappedCell(NamedTuple):
    cell: Cell

    def value(self) -> Any:
        return self.cell.value

    def is_none(self) -> bool:
        return is_none(self.cell)

    def is_number(self) -> bool:
        return is_number(self.cell)

    def ensure_int(self) -> int:
        return ensure_int(self.cell)

    # def ensure_int_or_none(self) -> int | None:
    def ensure_int_or_none(self) -> Union[int, None]:
        return ensure_int_or_none(self.cell)

    def non_empty_string(self, float_to_str: bool = False) -> str:
        return non_empty_string(self.cell, float_to_str)

    # def str_parse_or_none(self) -> str | None:
    def str_parse_or_none(self) -> Union[str, None]:
        return str_parse_or_none(self.cell)

    def float_parse(self) -> float:
        return float_parse(self.cell)

    # def float_parse_or_none(self) -> float | None:
    def float_parse_or_none(self) -> Union[float, None]:
        return float_parse_or_none(self.cell)

    def parse_date(self) -> NaiveDateTime:
        return parse_date(self.cell)


class WrappedWorksheet:
    _sheet: Worksheet
    _ncols: int

    def __init__(self, sheet: Worksheet) -> None:
        self._sheet = sheet
        self._ncols = _ncols_without_trailing_empty(sheet)

    def nrows(self) -> int:
        return self._sheet.max_row

    def ncols(self) -> int:
        return self._ncols

    def row(self, row_zero_indexed: int) -> Sequence[WrappedCell]:
        return [
            WrappedCell(cell)
            for cell in get_row(self._sheet, row_zero_indexed, self._ncols)
        ]

    def iter_rows(
        self, min_row_zero_indexed: int, max_row_zero_indexed: int
    ) -> Iterable[Sequence[WrappedCell]]:

        for row in iter_rows(
            self._sheet, min_row_zero_indexed, max_row_zero_indexed, self._ncols
        ):
            yield [WrappedCell(cell) for cell in row]


def _ncols_without_trailing_empty(worksheet: Worksheet) -> int:
    """Helper for difference in how openpyxl and xlrd treat trailing empty columns.

    xlrd strips trailing empty columns from the worksheet. openpyxl does not.

    Determine here the number of columns excluding trailing empty columns, so that
    we can emulate the behavior of xlrd.

    Optimized to return after reading the first row when there are no trailing
    empty columns.
    """
    sheet_max_ncols_non_empty = 0
    for row in worksheet.iter_rows():
        row_ncols_non_empty = 0
        for cell_idx_zero_indexed in range(len(row) - 1, -1, -1):
            cell = row[cell_idx_zero_indexed]
            if cell.value is not None:
                row_ncols_non_empty = cell_idx_zero_indexed + 1
                break

        sheet_max_ncols_non_empty = max(sheet_max_ncols_non_empty, row_ncols_non_empty)

        if sheet_max_ncols_non_empty == worksheet.max_column:
            break

    return sheet_max_ncols_non_empty


class WrappedWorkbook(NamedTuple):
    workbook: Workbook

    def sheet_names(self) -> Sequence[str]:
        return self.workbook.sheetnames

    def sheet_by_name(self, name: str) -> WrappedWorksheet:
        return WrappedWorksheet(self.workbook[name])
