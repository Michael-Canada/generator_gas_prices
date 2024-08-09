import io
from typing import IO, Iterable, Protocol, Sequence, Any, Union

import openpyxl
import xlrd

# from sourcing.reading.commons import opxl_commons, xlrd_commons
import opxl_commons
import xlrd_commons
from sourcing.utils.date_utils import NaiveDateTime


class Cell(Protocol):
    def value(self) -> Any:
        ...

    def is_none(self) -> bool:
        ...

    def is_number(self) -> bool:
        ...

    def ensure_int(self) -> int:
        ...

    def ensure_int_or_none(self) -> Union[int, None]:
        ...

    def non_empty_string(self, float_to_str: bool = False) -> str:
        ...

    def str_parse_or_none(self) -> Union[str, None]:
        ...

    def float_parse(self) -> float:
        ...

    def float_parse_or_none(self) -> Union[float, None]:
        ...

    def parse_date(self) -> NaiveDateTime:
        ...


class Worksheet(Protocol):
    def nrows(self) -> int:
        ...

    def ncols(self) -> int:
        ...

    def row(self, row_zero_indexed: int) -> Sequence[Cell]:
        ...

    def iter_rows(
        self, min_row_zero_indexed: int, max_row_zero_indexed: int
    ) -> Iterable[Sequence[Cell]]:
        ...


class Workbook(Protocol):
    def sheet_names(self) -> Sequence[str]:
        ...

    def sheet_by_name(self, name: str) -> Worksheet:
        ...


def open_workbook(contents: IO[bytes]) -> Workbook:
    contents: bytes = contents.read()

    file_format = xlrd.inspect_format(content=contents)

    if file_format == "xls":
        workbook = xlrd.open_workbook(file_contents=contents)
        return xlrd_commons.WrappedWorkbook(workbook)
    elif file_format == "xlsx":
        workbook = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
        return opxl_commons.WrappedWorkbook(workbook)
    else:
        raise ValueError(f"unexpected Excel file type: {file_format}")
